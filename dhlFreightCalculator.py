#!/usr/bin/env python
"""A freight calculator that uses a Tkinter GUI and pulls rates from a DHL rate .xlsx file"""
import os
import sys
import openpyxl
import tkinter as tk
import warnings

__author__ = 'Andrew Rice'
__copyright__ = 'Copyright 2019, Andrew Rice'
__credits__ = []
__license__ = ''
__version__ = '2.0'
__maintainer__ = 'Andrew Rice'
__email__ = ''
__status__ = 'Complete'


class Application(tk.Frame):
    def __init__(self, master = None):
        super().__init__(master)
        self.master = master

        # Variables for the DHL rate worksheet
        self.wb = None
        self.ratesWS = None

        self.countries = {'China', 'Taiwan', 'Hong Kong', 'Thailand'}
        self.country = 'China'    # The country chosen by the user to generate a landed freight quote from
        self.countryColumn = None    # The corresponding column of export country on the DHL .XLSX sheet
        self.countryChoices = None
        self.countryMenu = None

        # Method is the basis the user decides to quote by - weight per piece or weight of entire shipment
        self.methods = {'Per Piece', 'Entire Shipment'}
        self.method = 'Per Piece'    # default to calculate freight on per piece basis
        self.methodChoices = None
        self.methodMenu = None

        self.pieceWeightEntry = None  # entry box for weight of 1 piece
        self.pieceWeight = tk.StringVar()
        self.qty1 = tk.StringVar()
        self.qty1Float = 0.0
        self.qtyEntry1 = None
        self.qty2 = tk.StringVar()
        self.qty2Float = 0.0
        self.qtyEntry2 = None
        self.qty3 = tk.StringVar()
        self.qty3Float = 0.0
        self.qtyEntry3 = None
        self.qty4 = tk.StringVar()
        self.qty4Float = 0.0
        self.qtyEntry4 = None
        self.qty5 = tk.StringVar()
        self.qty5Float = 0.0
        self.qtyEntry5 = None

        self.grossWeight = tk.StringVar()    # gross weight of entire shipment
        self.grossWeightEntry = None    # entry box for weight of entire shipment

        self.headerFrame = None
        self.headerLabel = None
        self.instructionFrame = None
        self.instructionLabel = None
        self.countryFrame = None
        self.weightFrame = None
        self.footerFrame = None

        self.rateLabel1 = None    # Always a text label saying 'The cost of this shipment is estimated at:'
        self.rateLabel2 = None
        self.rateLabel3 = None
        self.rateLabel4 = None
        self.rateLabel5 = None
        self.rateLabel6 = None

        # Buttons
        self.calculateButton = None
        self.clearButton = None
        self.quitButton = None

        # Weight values and lists
        self.weightFloat = 0.0
        self.adjustedWeight = 0
        self.quotedWeight = 0    # final weight when quoting by entire shipment
        self.bufferDict = {    # Key = weight; Value = buffer to add
            0: 0,
            5: 1,
            10: 2,
            25: 2.5,
            50: 5,
            75: 8,
            100: 12,
            150: 17,
            200: 20,
            500: 35
        }
        self.freight_cost_raw = []
        self.freight_cost_adjusted = []
        self.freight_cost_buffed = []
        self.freight_cost_final = []  # final weight when quoting by individual quantities

        # Initializing functions
        self.pack()
        self.create_widgets()
        self.initialize_rates()

    def create_widgets(self):
        """Creates all widgets"""
        # create the Header Frame
        self.headerFrame = tk.Frame(self.master, width = 600, height = 50, bd = 1, highlightbackground = 'black',
                                    highlightcolor = 'black', highlightthickness = 0)
        self.headerFrame.place(x = 0, y = 0)
        self.headerFrame.pack_propagate(False)
        self.headerLabel = tk.Label(self.headerFrame, text = 'DHL Freight Calculator', fg = 'red',
                                    font = ('calibri', 16))
        self.headerLabel.place(x = 300, y = 37.5, anchor = 'center')

        # Create an instruction box underneath the header
        self.instructionFrame = tk.Frame(self.master, width = 600, height = 150, bd = 0, highlightbackground = 'black',
                                         highlightcolor = 'black', highlightthickness = 0)
        self.instructionFrame.place(x=0, y=50)
        self.instructionFrame.pack_propagate(False)
        self.instructionLabel = tk.Label(self.instructionFrame, text = 'This is a comprehensive freight quote '
                                                                       'generator that uses our custom DHL rates to '
                                                                       'generate a freight quote based on multiple '
                                                                       'quantities and country of origin. Please '
                                                                       'select a country from the drop down menu, '
                                                                       'enter the quantities you would like to quote, '
                                                                       'and the weight of 1 piece to generate a quote. '
                                                                       'Alternatively, if you know the weight of the '
                                                                       'entire shipment, please select that option '
                                                                       'and enter the gross weight to generate your '
                                                                       'quote.',
                                         height = 125, wraplength = 450, justify = 'center', font = ('calibri', 10))
        self.instructionLabel.place(x = 300, y = 75, anchor = 'center')

        # Create Frame where user is able to select country and method
        self.countryFrame = tk.Frame(self.master, width = 600, height = 50, bd = 1, highlightbackground = 'black',
                                     highlightcolor = 'black', highlightthickness = 0)
        self.countryFrame.place(x = 0, y = 200)

        # TODO: change menu and choices for country and method from grid to place
        # Create a drop down for users to choose which country they are exporting from
        self.countryChoices = tk.StringVar(self.master)
        self.countryChoices.set('China')
        self.countryMenu = tk.OptionMenu(self.countryFrame, self.countryChoices, *self.countries)
        tk.Label(self.countryFrame, text = 'Select a Country:').grid(row = 1, column = 1, padx = 20)
        self.countryMenu.grid(row = 1, column = 2)
        self.countryChoices.trace('w', self.change_country_dropdown)

        # Create a drop down for users to choose whether they are quoting weight per piece or entire shipment
        self.methodChoices = tk.StringVar(self.master)
        self.methodChoices.set('Per Piece')
        self.methodMenu = tk.OptionMenu(self.countryFrame, self.methodChoices, *self.methods)
        tk.Label(self.countryFrame, text = 'How are you calculating freight:').grid(row = 1, column = 3, padx = 20)
        self.methodMenu.grid(row = 1, column = 4)
        self.methodChoices.trace('w', self.change_method_dropdown)

        # Initialize weight frame
        self.set_weight_frame()

        # Create Footer Frame
        self.footerFrame = tk.Frame(self.master, width = 600, height = 50, bd = 1, highlightbackground = 'black',
                                    highlightcolor = 'black', highlightthickness = 0)
        self.footerFrame.place(x = 0, y = 700)

        # Create calculate button using generate_report script
        self.calculateButton = tk.Button(self.footerFrame, text = 'Calculate', command = self.generate_report,
                                         cursor = 'hand2')
        self.calculateButton.place(x = 225, y = 10, height = 30, width = 150)

        # Create a 'clear fields' button that wipes the form clean and allows users to run another quote
        self.clearButton = tk.Button(self.footerFrame, text = 'Clear Fields', command = self.clear_fields,
                                     cursor = 'hand2')
        self.clearButton.place(x = 25, y = 10, height = 30, width = 75)

        # Create a quit button
        self.quitButton = tk.Button(self.footerFrame, text = 'Quit', command = self.quit_application, cursor = 'hand2')
        self.quitButton.place(x = 500, y = 10, height = 30, width = 75)

        # TODO: Create label with program & copyright info

    def resourcePath(self, relativePath):
        """ Get absolute path to resource, works for dev and for PyInstaller """
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            basePath = sys._MEIPASS
        except Exception:
            basePath = os.path.abspath(".")

        return os.path.join(basePath, relativePath)

    def initialize_rates(self):
        """Initialize freight rates from DHL excel workbook"""
        with warnings.catch_warnings():  # .wmf image in the excel file causes a warning
            warnings.simplefilter('ignore')
            self.wb = openpyxl.load_workbook(self.resourcePath('./dhlRates.xlsx'), data_only = True)
            self.ratesWS = self.wb['US Import Rates']  # import rates

    def change_country_dropdown(self, *args):
        """Sets the country variable when users select a different country from the dropdown menu"""
        self.country = str(self.countryChoices.get())

    def change_method_dropdown(self, *args):
        """Change the dropdown menu for quote method - calls set_weight_frame"""
        self.clear_fields()
        self.method = str(self.methodChoices.get())
        self.weightFrame.destroy()

        # reset weight frame
        self.set_weight_frame()

    def set_weight_frame(self):
        """Set the weight frame that allows users to input quantities and weights, refreshes based on method"""
        self.weightFrame = tk.Frame(self.master, width = 600, height = 450, bd = 1, highlightbackground = 'black',
                                    highlightcolor = 'black', highlightthickness = 0)
        self.weightFrame.place(x = 0, y = 250)
        if self.method == 'Per Piece':
            tk.Label(self.weightFrame, text = 'Weight of 1 piece in grams (g):').place(x = 150, y = 15)
            self.pieceWeightEntry = tk.Entry(self.weightFrame, textvariable = self.pieceWeight)
            self.pieceWeightEntry.place(x = 325, y = 15)
            tk.Label(self.weightFrame, text = "Quantity 1:").place(x = 75, y= 75)
            tk.Label(self.weightFrame, text = "Quantity 2:").place(x = 75, y = 125)
            tk.Label(self.weightFrame, text = "Quantity 3:").place(x = 75, y = 175)
            tk.Label(self.weightFrame, text = "Quantity 4:").place(x = 75, y = 225)
            tk.Label(self.weightFrame, text = "Quantity 5:").place(x = 75, y = 275)
            self.qtyEntry1 = tk.Entry(self.weightFrame, textvariable = self.qty1)
            self.qtyEntry2 = tk.Entry(self.weightFrame, textvariable = self.qty2)
            self.qtyEntry3 = tk.Entry(self.weightFrame, textvariable = self.qty3)
            self.qtyEntry4 = tk.Entry(self.weightFrame, textvariable = self.qty4)
            self.qtyEntry5 = tk.Entry(self.weightFrame, textvariable = self.qty5)
            self.qtyEntry1.place(x = 150, y = 75)
            self.qtyEntry2.place(x = 150, y = 125)
            self.qtyEntry3.place(x = 150, y = 175)
            self.qtyEntry4.place(x = 150, y = 225)
            self.qtyEntry5.place(x = 150, y = 275)

        else:
            # Create a box for user to enter weight entire shipment
            tk.Label(self.weightFrame, text = 'Gross Weight of Shipment in Kilograms (kg):').place(x = 75, y = 15)
            # self.grossWeight = tk.StringVar()
            self.grossWeightEntry = tk.Entry(self.weightFrame, textvariable = self.grossWeight)
            self.grossWeightEntry.place(x = 325, y = 15)
            # TODO: Bind the entry field to the enter key, so when the user hits enter it runs the script
            # self.grossWeightEntry.bind('<Return>', self.calculateButton.invoke())

    def generate_report(self):
        """Grab the weight and quantity value and generate a value or list of values determined by method"""
        try:
            if self.method == 'Per Piece':
                try:
                    self.weightFloat = float(self.pieceWeight.get())/1000    # set weight of 1 pc in kilograms
                except ValueError:
                    pass
                try:
                    self.qty1Float = float(self.qty1.get())
                except ValueError:
                    self.qty1Float = 0
                try:
                    self.qty2Float = float(self.qty2.get())
                except ValueError:
                    self.qty2Float = 0
                try:
                    self.qty3Float = float(self.qty3.get())
                except ValueError:
                    self.qty3Float = 0
                try:
                    self.qty4Float = float(self.qty4.get())
                except ValueError:
                    self.qty4Float = 0
                try:
                    self.qty5Float = float(self.qty5.get())
                except ValueError:
                    self.qty5Float = 0

                self.freight_cost_raw = [float(self.qty1Float * self.weightFloat),
                                         float(self.qty2Float * self.weightFloat),
                                         float(self.qty3Float * self.weightFloat),
                                         float(self.qty4Float * self.weightFloat),
                                         float(self.qty5Float * self.weightFloat)]
                self.adjust()
                self.buffer()
                self.set_country_column()
                self.generate_cost()
                self.display_rates()

                # Clear lists to allow user to calculate again
                self.weightFloat = 0
                self.freight_cost_raw = []
                self.freight_cost_adjusted = []
                self.freight_cost_buffed = []
                self.freight_cost_final = []
            else:
                try:
                    self.freight_cost_raw = [float(self.grossWeight.get())]
                except ValueError:
                    self.freight_cost_raw = 0
                try:
                    self.weightFloat = float(self.grossWeight.get())
                except ValueError:
                    self.weightFloat = 0
                self.adjust()
                self.buffer()
                self.set_country_column()
                self.generate_cost()
                self.display_rates()

                self.weightFloat = 0
                self.freight_cost_raw = 0
        except AttributeError as err:
            print(err)
            pass

    def adjust(self):
        """Rounds the calculated weight up to the nearest 1/2 KG"""
        if self.method == 'Per Piece':
            for i in self.freight_cost_raw:
                dec = float(i) - int(i)
                if (dec < .5) and (dec > 0):
                    dec = .5
                elif dec >= .5:
                    dec = 1
                else:
                    dec = 0
                self.freight_cost_adjusted.append(int(i) + dec)
        else:
            dec = float(self.weightFloat) - int(self.weightFloat)
            if (dec < .5) and (dec > 0):
                dec = .5
            elif dec >= .5:
                dec = 1
            else:
                dec = 0
            self.adjustedWeight = int(self.weightFloat) + dec

    def buffer(self):  # Add a buffer to the calculated weights
        """Adds a buffer to the weight to give cushion when estimating"""
        buffer = 0
        if self.method == 'Per Piece':
            for i in self.freight_cost_adjusted:
                for wt, bf in self.bufferDict.items():
                    if i <= wt:
                        buffer = bf
                        break
                    else:
                        buffer = 40
                self.freight_cost_buffed.append(float(i) + buffer)
        else:
            for wt, bf in self.bufferDict.items():
                if self.adjustedWeight <= wt:
                    buffer = bf
                    break
                else:
                    buffer = 40
            self.adjustedWeight += buffer

    def set_country_column(self):
        """Set the country column variable when the user changes the active country"""
        if self.country == 'China':
            self.countryColumn = 9
        elif self.country == 'Taiwan' or self.country == 'Hong Kong':
            self.countryColumn = 10
        elif self.country == 'Thailand':
            self.countryColumn = 11

    def generate_cost(self):
        """Iterate through the weight list to calculate a cost using the DHL rate sheet"""
        if self.method == 'Per Piece':
            for x in self.freight_cost_buffed:
                wt = float(x)
                col = self.countryColumn
                if wt <= 0:
                    wt = 0
                elif wt <= 70:
                    for i in range(15, 155, 1):
                        if wt == float(self.ratesWS.cell(row = i, column = 1).value):
                            wt = float(self.ratesWS.cell(row = i, column = col).value)
                elif wt <= 150:
                    wt *= float(self.ratesWS.cell(row = 158, column = col).value)
                elif wt <= 300:
                    wt *= float(self.ratesWS.cell(row = 159, column = col).value)
                elif wt <= 999:
                    wt *= float(self.ratesWS.cell(row = 160, column = col).value)
                else:
                    wt *= float(self.ratesWS.cell(row = 161, column = col).value)
                self.freight_cost_final.append(wt)
        else:
            wt = float(self.adjustedWeight)
            col = self.countryColumn
            if wt <= 70:
                for i in range(15, 155, 1):
                    if wt == float(self.ratesWS.cell(row = i, column = 1).value):
                        wt = float(self.ratesWS.cell(row = i, column = col).value)
            elif wt <= 150:
                wt *= float(self.ratesWS.cell(row = 158, column = col).value)
            elif wt <= 300:
                wt *= float(self.ratesWS.cell(row = 159, column = col).value)
            elif wt <= 999:
                wt *= float(self.ratesWS.cell(row = 160, column = col).value)
            else:
                wt *= float(self.ratesWS.cell(row = 161, column = col).value)
            self.quotedWeight = wt

    def display_rates(self):
        """Open a second tkinter window that displays the quoted rate calculated by the program"""
        try:
            self.rateLabel1.destroy()
            self.rateLabel2.destroy()
            self.rateLabel3.destroy()
            self.rateLabel4.destroy()
            self.rateLabel5.destroy()
            self.rateLabel6.destroy()
        except AttributeError:
            pass

        self.rateLabel1 = tk.Label(self.weightFrame, text = 'The costs of this shipment are estimated at:')

        if self.method == 'Per Piece':
            self.rateLabel2 = tk.Label(self.weightFrame, text = '${:,.2f}'.format(self.freight_cost_final[0]),
                                       fg = 'green', font = 14)
            self.rateLabel3 = tk.Label(self.weightFrame, text = '${:,.2f}'.format(self.freight_cost_final[1]),
                                       fg = 'green', font = 14)
            self.rateLabel4 = tk.Label(self.weightFrame, text = '${:,.2f}'.format(self.freight_cost_final[2]),
                                       fg = 'green', font = 14)
            self.rateLabel5 = tk.Label(self.weightFrame, text = '${:,.2f}'.format(self.freight_cost_final[3]),
                                       fg = 'green', font = 14)
            self.rateLabel6 = tk.Label(self.weightFrame, text = '${:,.2f}'.format(self.freight_cost_final[4]),
                                       fg = 'green', font = 14)

            if self.freight_cost_final[0] > 0:
                self.rateLabel2.place(x = 300, y = 70)
            if self.freight_cost_final[1] > 0:
                self.rateLabel3.place(x = 300, y = 120)
            if self.freight_cost_final[2] > 0:
                self.rateLabel4.place(x = 300, y = 170)
            if self.freight_cost_final[3] > 0:
                self.rateLabel5.place(x = 300, y = 220)
            if self.freight_cost_final[4] > 0:
                self.rateLabel6.place(x = 300, y = 270)
        else:
            self.rateLabel2 = tk.Label(self.weightFrame, text = '${:,.2f}'.format(self.quotedWeight),
                                       fg = 'green', font = 16)
            if self.quotedWeight > 0:
                self.rateLabel1.place(x = 300, y = 100, anchor = 'center')
                self.rateLabel2.place(x = 300, y = 130, anchor = 'center')

    def clear_fields(self):
        """Clear all fields in the weightFrame"""
        try:
            if self.method == 'Per Piece':
                self.pieceWeightEntry.delete(0, 'end')
                self.qtyEntry1.delete(0, 'end')
                self.qtyEntry2.delete(0, 'end')
                self.qtyEntry3.delete(0, 'end')
                self.qtyEntry4.delete(0, 'end')
                self.qtyEntry5.delete(0, 'end')
                self.rateLabel1.destroy()
                self.rateLabel2.destroy()
                self.rateLabel3.destroy()
                self.rateLabel4.destroy()
                self.rateLabel5.destroy()
                self.rateLabel6.destroy()
            else:
                self.grossWeightEntry.delete(0, 'end')
                self.rateLabel1.destroy()
                self.rateLabel2.destroy()
        except AttributeError:
            pass

    def quit_application(self):
        """Quit the application"""
        self.wb.close()
        self.master.destroy()


if __name__ == '__main__':
    calcApp = tk.Tk()
    calcApp.title('DHL Freight Calculator')
    calcApp.geometry('600x750')
    calcApp.resizable(0, 0)
    app = Application(master = calcApp)
    app.mainloop()
    app.wb.close()
